"""
Hulpfuncties voor de applicatie
"""
import os
import json
import logging
import uuid
import calendar
import re
from functools import wraps
from datetime import datetime, timedelta
from flask import request, abort, current_app
from flask_login import current_user, login_required


def format_currency(value):
    """
    Formatteer een valutawaarde
    
    Args:
        value: De numerieke waarde
        
    Returns:
        str: De geformatteerde waarde met 2 decimalen
    """
    return '{:.2f}'.format(value)


def format_decimal(value, decimal_places=2):
    """
    Formatteer een decimale waarde
    
    Args:
        value: De numerieke waarde
        decimal_places: Aantal decimalen (standaard 2)
        
    Returns:
        str: De geformatteerde waarde met het opgegeven aantal decimalen
    """
    format_str = '{:.' + str(decimal_places) + 'f}'
    return format_str.format(value)


def generate_token():
    """
    Genereer een uniek token
    
    Returns:
        str: Een unieke string token
    """
    return str(uuid.uuid4())


def add_months(sourcedate, months):
    """
    Voeg een aantal maanden toe aan een datum
    
    Args:
        sourcedate: De oorspronkelijke datum
        months: Het aantal toe te voegen maanden
        
    Returns:
        datetime: De resulterende datum
    """
    month = sourcedate.month - 1 + months
    year = sourcedate.year + month // 12
    month = month % 12 + 1
    day = min(sourcedate.day, [31, 29 if is_leap_year(year) else 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31][month - 1])
    return sourcedate.replace(year=year, month=month, day=day)


def is_leap_year(year):
    """
    Controleer of een jaar een schrikkeljaar is
    
    Args:
        year: Het jaar om te controleren
        
    Returns:
        bool: True als het een schrikkeljaar is, anders False
    """
    return year % 4 == 0 and (year % 100 != 0 or year % 400 == 0)


def encrypt_data(data):
    """
    Versleutel gevoelige gegevens
    
    Args:
        data: De gegevens om te versleutelen
        
    Returns:
        str: De versleutelde gegevens
    """
    # Eenvoudige implementatie: encodeer naar JSON en dan naar base64
    # In een productie-omgeving zou dit vervangen moeten worden door echte encryptie
    import base64
    return base64.b64encode(json.dumps(data).encode('utf-8')).decode('utf-8')


def decrypt_data(encrypted_data):
    """
    Ontsleutel versleutelde gegevens
    
    Args:
        encrypted_data: De versleutelde gegevens
        
    Returns:
        dict: De ontsleutelde gegevens of None bij fouten
    """
    # Eenvoudige implementatie: decodeer van base64 en dan van JSON
    # In een productie-omgeving zou dit vervangen moeten worden door echte decryptie
    try:
        import base64
        return json.loads(base64.b64decode(encrypted_data).decode('utf-8'))
    except Exception as e:
        logging.error(f"Fout bij ontsleutelen gegevens: {str(e)}")
        return None


def make_public_url(path):
    """
    Genereer een openbare URL voor een bestandspad
    
    Args:
        path: Het bestandspad
        
    Returns:
        str: De openbare URL
    """
    base_url = os.environ.get('BASE_URL', 'http://localhost:5000')
    if path.startswith('/'):
        path = path[1:]
    return f"{base_url}/{path}"


def generate_password():
    """
    Genereer een willekeurig wachtwoord
    
    Returns:
        str: Een willekeurig gegenereerd wachtwoord
    """
    import random
    import string
    
    # Genereer een wachtwoord van 12 tekens met minimaal 1 hoofdletter, 1 cijfer en 1 speciaal teken
    chars = string.ascii_lowercase
    chars_upper = string.ascii_uppercase
    chars_digits = string.digits
    chars_special = '!@#$%^&*'
    
    password = [
        random.choice(chars_upper),
        random.choice(chars_digits),
        random.choice(chars_special)
    ]
    
    # Vul aan tot 12 tekens
    for _ in range(9):
        password.append(random.choice(chars + chars_upper + chars_digits + chars_special))
    
    # Shuffle en maak een string
    random.shuffle(password)
    return ''.join(password)


def validate_email_domain(email, domain):
    """
    Controleer of een e-mailadres overeenkomt met een bepaald domein
    
    Args:
        email: Het e-mailadres om te controleren
        domain: Het domein om te vergelijken
        
    Returns:
        bool: True als het e-mailadres overeenkomt met het domein, anders False
    """
    if not domain:
        return True  # Als er geen domein is opgegeven, accepteer dan elk e-mailadres
    
    # Haal het domeingedeelte van het e-mailadres
    email_parts = email.split('@')
    if len(email_parts) != 2:
        return False
    
    email_domain = email_parts[1].lower()
    domain = domain.lower()
    
    # Controleer of het domein overeenkomt of een subdomein is
    return email_domain == domain or email_domain.endswith('.' + domain)


def is_safe_url(url):
    """
    Controleer of een URL veilig is om naar door te verwijzen
    
    Args:
        url: De URL om te controleren
        
    Returns:
        bool: True als de URL veilig is, anders False
    """
    from urllib.parse import urlparse, urljoin
    from flask import request
    
    ref_url = urlparse(request.host_url)
    test_url = urlparse(urljoin(request.host_url, url))
    
    # Controleer of de URL naar dezelfde site gaat
    return test_url.scheme in ('http', 'https') and ref_url.netloc == test_url.netloc


def generate_pdf_invoice(invoice_data):
    """
    Genereer een PDF-factuur op basis van factuurgegevens
    
    Args:
        invoice_data: Dict met factuurgegevens
        
    Returns:
        bytes: PDF-inhoud als bytes
    """
    # Dit is een placeholder functie die later geïmplementeerd zal worden
    # met een echte PDF-generatie bibliotheek zoals WeasyPrint of ReportLab
    return b"PDF CONTENT"


def export_to_excel(data, headers=None):
    """
    Exporteer gegevens naar Excel-formaat
    
    Args:
        data: Lijst van rijen met gegevens
        headers: Kolomkoppen (optioneel)
        
    Returns:
        bytes: Excel-bestand als bytes
    """
    import io
    import openpyxl
    from openpyxl import Workbook
    
    wb = Workbook()
    ws = wb.active
    
    # Voeg headers toe
    if headers:
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)
    
    # Voeg data toe
    row_num = 2 if headers else 1
    for row in data:
        for col_num, cell_value in enumerate(row, 1):
            ws.cell(row=row_num, column=col_num, value=cell_value)
        row_num += 1
    
    # Sla op naar bytes
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return excel_file.getvalue()


def export_to_csv(data, headers=None, delimiter=','):
    """
    Exporteer gegevens naar CSV-formaat
    
    Args:
        data: Lijst van rijen met gegevens
        headers: Kolomkoppen (optioneel)
        delimiter: Scheidingsteken (standaard komma)
        
    Returns:
        str: CSV-inhoud
    """
    import io
    import csv
    
    output = io.StringIO()
    writer = csv.writer(output, delimiter=delimiter, quoting=csv.QUOTE_MINIMAL)
    
    if headers:
        writer.writerow(headers)
    
    for row in data:
        writer.writerow(row)
    
    return output.getvalue()


def get_vat_rates():
    """
    Haal de BTW-tarieven op
    
    Returns:
        list: Lijst van BTW-tarieven voor België en Nederland
    """
    # BTW-tarieven voor België en Nederland
    return [
        {'code': 'NL_HIGH', 'rate': 21.0, 'description': 'Nederland - Hoog (21%)'},
        {'code': 'NL_LOW', 'rate': 9.0, 'description': 'Nederland - Laag (9%)'},
        {'code': 'NL_ZERO', 'rate': 0.0, 'description': 'Nederland - Nultarief (0%)'},
        {'code': 'BE_HIGH', 'rate': 21.0, 'description': 'België - Hoog (21%)'},
        {'code': 'BE_MEDIUM', 'rate': 12.0, 'description': 'België - Middel (12%)'},
        {'code': 'BE_LOW', 'rate': 6.0, 'description': 'België - Laag (6%)'},
        {'code': 'BE_ZERO', 'rate': 0.0, 'description': 'België - Nultarief (0%)'}
    ]


def date_to_quarter(date):
    """
    Converteer een datum naar een kwartaalaanduiding
    
    Args:
        date: Een datetime object
        
    Returns:
        str: Kwartaal aanduiding (bijv. "Q1 2023")
    """
    quarter = (date.month - 1) // 3 + 1
    return f"Q{quarter} {date.year}"


def get_quarters(start_year, end_year=None):
    """
    Genereer een lijst van kwartalen voor rapportage
    
    Args:
        start_year: Het startjaar
        end_year: Het eindjaar (optioneel, standaard is het startjaar)
        
    Returns:
        list: Lijst van kwartalen (bijv. ["Q1 2022", "Q2 2022", ...])
    """
    if not end_year:
        end_year = start_year
    
    quarters = []
    for year in range(start_year, end_year + 1):
        for quarter in range(1, 5):
            quarters.append(f"Q{quarter} {year}")
    
    return quarters


def get_months(start_year, start_month=1, end_year=None, end_month=12):
    """
    Genereer een lijst van maanden voor rapportage
    
    Args:
        start_year: Het startjaar
        start_month: De startmaand (standaard 1)
        end_year: Het eindjaar (optioneel, standaard is het startjaar)
        end_month: De eindmaand (standaard 12)
        
    Returns:
        list: Lijst van maanden (bijv. ["Jan 2022", "Feb 2022", ...])
    """
    if not end_year:
        end_year = start_year
    
    months = []
    month_names = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    
    for year in range(start_year, end_year + 1):
        for month in range(1, 13):
            if (year == start_year and month < start_month) or (year == end_year and month > end_month):
                continue
            months.append(f"{month_names[month-1]} {year}")
    
    return months


def get_years(start_year, end_year=None):
    """
    Genereer een lijst van jaren voor rapportage
    
    Args:
        start_year: Het startjaar
        end_year: Het eindjaar (optioneel, standaard is het startjaar)
        
    Returns:
        list: Lijst van jaren (bijv. [2021, 2022, 2023])
    """
    if not end_year:
        end_year = start_year
    
    return list(range(start_year, end_year + 1))


def save_uploaded_file(file, directory, filename=None):
    """
    Sla een geüpload bestand op
    
    Args:
        file: Het bestandsobject
        directory: De map waarin het bestand moet worden opgeslagen
        filename: De bestandsnaam (optioneel, standaard wordt de originele bestandsnaam gebruikt)
        
    Returns:
        str: Het pad naar het opgeslagen bestand
    """
    import os
    
    # Maak de map aan als deze nog niet bestaat
    os.makedirs(directory, exist_ok=True)
    
    # Bepaal bestandsnaam
    if not filename:
        filename = file.filename
    
    # Sla het bestand op
    file_path = os.path.join(directory, filename)
    file.save(file_path)
    
    return file_path


def allowed_file(filename, allowed_extensions):
    """
    Controleer of een bestandsnaam een toegestane extensie heeft
    
    Args:
        filename: De bestandsnaam om te controleren
        allowed_extensions: Lijst van toegestane extensies (bijv. ['pdf', 'png', 'jpg'])
        
    Returns:
        bool: True als de extensie is toegestaan, anders False
    """
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in allowed_extensions


# Logging helpers
def highlight_log_line(line, search_term):
    """
    Markeer een zoekterm in een log regel met HTML highlighting
    
    Args:
        line: De log regel
        search_term: De te markeren zoekterm
        
    Returns:
        str: De gemarkeerde regel
    """
    if not search_term:
        return line
        
    # Simpele markering met span-tag
    marked_line = line.replace(
        search_term, 
        f'<span class="highlight">{search_term}</span>'
    )
    return marked_line

def get_log_context(filepath, line_number, context_lines=5):
    """
    Haal context rond een specifieke regel in een logbestand op
    
    Args:
        filepath: Pad naar het logbestand
        line_number: Regelnummer (beginnend bij 1)
        context_lines: Aantal regels context voor en na (standaard 5)
        
    Returns:
        list: Lijst met context regels
    """
    if not os.path.exists(filepath):
        return []
        
    context = []
    start_line = max(1, line_number - context_lines)
    end_line = line_number + context_lines
    
    with open(filepath, 'r', encoding='utf-8', errors='replace') as f:
        for i, line in enumerate(f, 1):
            if start_line <= i <= end_line:
                # Markeer de hoofdregel
                if i == line_number:
                    context.append((i, line.strip(), True))
                else:
                    context.append((i, line.strip(), False))
                    
            if i > end_line:
                break
                
    return context

def analyze_log_errors(log_file, max_entries=100):
    """
    Analyseer logbestand voor veelvoorkomende fouten
    
    Args:
        log_file: Pad naar het logbestand
        max_entries: Maximum aantal fouten om te analyseren
        
    Returns:
        dict: Analyse van fouten (frequentie, voorbeelden)
    """
    if not os.path.exists(log_file):
        return {}
        
    error_patterns = {}
    
    with open(log_file, 'r', encoding='utf-8', errors='replace') as f:
        for line in f:
            if 'ERROR' not in line and 'CRITICAL' not in line:
                continue
                
            # Probeer een foutpatroon te identificeren
            patterns = [
                r'Error: (.*?)( at |$)',
                r'Exception: (.*?)( at |$)',
                r'ERROR .* (.*?)( in |$)',
                r'CRITICAL .* (.*?)( in |$)'
            ]
            
            error_text = None
            for pattern in patterns:
                match = re.search(pattern, line)
                if match:
                    error_text = match.group(1).strip()
                    break
            
            if not error_text:
                error_text = "Unknown error"
                
            # Houd foutfrequentie bij
            if error_text in error_patterns:
                error_patterns[error_text]['count'] += 1
                if len(error_patterns[error_text]['examples']) < 3:  # Max 3 voorbeelden
                    error_patterns[error_text]['examples'].append(line.strip())
            else:
                error_patterns[error_text] = {
                    'count': 1,
                    'examples': [line.strip()]
                }
                
            # Stop als we genoeg fouten hebben
            if len(error_patterns) >= max_entries:
                break
                
    # Sorteer op frequentie
    sorted_patterns = dict(sorted(
        error_patterns.items(), 
        key=lambda item: item[1]['count'], 
        reverse=True
    ))
    
    return sorted_patterns

# Admin vereist decorator (voor log viewer)
def admin_required(f):
    """
    Decorator die controleert of een gebruiker admin-rechten heeft
    
    Args:
        f: De functie om te decoreren
        
    Returns:
        function: Gedecoreerde functie
    """
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated:
            return abort(401)  # Unauthorized
        
        if not current_user.is_admin:
            return abort(403)  # Forbidden
            
        return f(*args, **kwargs)
    return decorated_function

# Gebruikersrechten functies
from functools import wraps
from flask import flash, redirect, url_for

def check_permission(permission_name):
    """
    Controleert of de ingelogde gebruiker een specifieke permissie heeft
    
    Args:
        permission_name: Naam van de permissie om te controleren (bijv. 'can_view_customers')
        
    Returns:
        bool: True als de gebruiker de permissie heeft, anders False
    """
    # Super admins hebben altijd alle rechten
    if current_user.is_super_admin:
        return True
    
    # Normale admins hebben standaard beheerrechten
    if current_user.is_admin and permission_name not in [
        'can_view_customers', 'can_add_customers', 'can_edit_customers', 'can_delete_customers',
        'can_view_invoices', 'can_add_invoices', 'can_edit_invoices', 'can_delete_invoices', 'can_upload_invoices',
        'can_view_reports', 'can_export_reports', 'can_generate_vat_report',
        'can_view_dashboard', 'can_manage_settings'
    ]:
        return True
    
    # Controleer gebruikersrechten
    permissions = current_user.permissions
    if not permissions:
        return False
    
    # Controleer specifieke permissie
    return getattr(permissions, permission_name, False)

def permission_required(permission_name):
    """
    Decorator die controleert of een gebruiker een bepaalde permissie heeft
    
    Args:
        permission_name: Naam van de permissie die vereist is
        
    Returns:
        function: Decorator functie
    """
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if not current_user.is_authenticated:
                flash('U moet ingelogd zijn om deze pagina te bekijken', 'danger')
                return redirect(url_for('login'))
            
            if not check_permission(permission_name):
                flash('U heeft geen toegang tot deze pagina', 'danger')
                return redirect(url_for('dashboard'))
            
            return f(*args, **kwargs)
        return decorated_function
    return decorator